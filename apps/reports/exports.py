"""تصدير التقرير المالي إلى PDF (reportlab) أو Excel (openpyxl).

الاستيرادات محمية داخل الدوال حتى لا يفشل التطبيق إن لم تُثبَّت المكتبات؛
تُرفع BusinessRuleError واضحة عند غياب المكتبة أو صيغة غير مدعومة.
"""
from decimal import Decimal
from io import BytesIO

from apps.core.exceptions import BusinessRuleError

ZERO = Decimal("0.00")


def _summary_rows(report):
    """صفوف الملخّص المالي (تسمية عربية، قيمة) بترتيب ثابت."""
    return [
        ("الهدف", report["target"]),
        ("التعهدات", report["pledged"]),
        ("المحصّل", report["collected"]),
        ("المتبقي", report["remaining"]),
        ("المصروفات", report["expenses"]),
        ("الرصيد", report["balance"]),
        ("قيمة العيني", report["inkind_value"]),
        ("عدد المساهمين", report["contributors"]),
        ("عدد الاشتراكات", report["subscriptions"]),
        ("الإنجاز المالي %", report["financial_progress"]),
        ("الإنجاز التنفيذي %", report["execution_progress"]),
    ]


def export_financial_pdf(report):
    """يُرجع (bytes, content_type, filename) لملف PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError as exc:  # pragma: no cover - depends on optional dep
        raise BusinessRuleError(
            "مكتبة reportlab غير مثبّتة لتصدير PDF.",
            code="export_backend_missing",
        ) from exc

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title="التقرير المالي",
    )
    styles = getSampleStyleSheet()
    elements = []

    project = report["project"]
    elements.append(
        Paragraph(f"Financial Report - {project['reference']}", styles["Title"])
    )
    elements.append(Paragraph(project["name"], styles["Heading2"]))
    elements.append(Spacer(1, 0.5 * cm))

    data = [["البند", "القيمة"]]
    for label, value in _summary_rows(report):
        data.append([label, str(value)])

    table = Table(data, colWidths=[8 * cm, 8 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    filename = f"financial_{project['reference']}.pdf"
    return pdf, "application/pdf", filename


def export_financial_excel(report):
    """يُرجع (bytes, content_type, filename) لملف Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - depends on optional dep
        raise BusinessRuleError(
            "مكتبة openpyxl غير مثبّتة لتصدير Excel.",
            code="export_backend_missing",
        ) from exc

    project = report["project"]
    wb = Workbook()

    # ورقة الملخّص.
    ws = wb.active
    ws.title = "الملخّص"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F766E")

    ws.append([f"التقرير المالي — {project['name']} ({project['reference']})"])
    ws.append([])
    ws.append(["البند", "القيمة"])
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
    for label, value in _summary_rows(report):
        ws.append([label, float(value) if isinstance(value, Decimal) else value])

    # ورقة الإيرادات حسب النوع.
    ws_rev = wb.create_sheet("الإيرادات حسب النوع")
    ws_rev.append(["النوع", "الإجمالي", "العدد"])
    for cell in ws_rev[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in report["revenues_by_type"]:
        ws_rev.append(
            [row["revenue_type"], float(row["total"] or 0), row["count"]]
        )

    # ورقة المصروفات حسب التصنيف.
    ws_exp = wb.create_sheet("المصروفات حسب التصنيف")
    ws_exp.append(["التصنيف", "الإجمالي", "العدد"])
    for cell in ws_exp[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in report["expenses_by_category"]:
        ws_exp.append(
            [row["category"] or "غير مصنّف", float(row["total"] or 0), row["count"]]
        )

    buffer = BytesIO()
    wb.save(buffer)
    data = buffer.getvalue()
    buffer.close()
    filename = f"financial_{project['reference']}.xlsx"
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return data, content_type, filename
